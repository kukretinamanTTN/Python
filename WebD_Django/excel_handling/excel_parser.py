from openpyxl import load_workbook

# wb = load_workbook("data.xlsx")
# sheet = wb.active
# for row in sheet.iter_rows(values_only=True):  
#     print(row)

with open("data.xlsx", "rb+") as f, open("uploaded.xlsx", "wb+") as destination:
    for chunk in f.chunks():  
            destination.write(chunk)
