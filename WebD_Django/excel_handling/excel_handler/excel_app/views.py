from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from .forms import UploadFileForm
from .models import PollQuestion

#validate Excel file
def validate_excel(file):
    if not file.name.endswith('.xlsx'):
        raise ValidationError("Only .xlsx files are allowed")
    if file.size > 5 * 1024 * 1024:  # 5MB limit
        raise ValidationError("File size exceeds limit")

#parse and store data from Excel
def handle_uploaded_file(file):
    try:
        validate_excel(file)

        wb = load_workbook(file)
        sheet = wb.active  
        for row in sheet.iter_rows(max_row=4, values_only=True):
            question_text = row[0]

            if not question_text:
                continue

            PollQuestion.objects.create(question_text=question_text)  #save to DB

    except ValidationError as e:
        return f"Validation Error: {e}"
    except Exception as e:
        return f"Failed to process file: {e}"

#handle file upload and parsing
def upload_file(request):
    form=[]
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            error_message = handle_uploaded_file(request.FILES['file'])
            if error_message:
                return render(request, "upload.html", {"form": form, "error": error_message})
            return HttpResponse("<h2>File uploaded to DB successfully!</h2>")
        form = UploadFileForm()
    
    return render(request, "upload.html", {"form": form})